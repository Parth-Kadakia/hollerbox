"""Best-effort text extraction from non-image attachments.

The `llm` step uses this so chat-attached PDFs / spreadsheets / CSVs
can be sent to providers that don't natively read those formats. We
classify by mime + extension, extract what we can, and return either:

- the original `Attachment` (no extraction; the provider handles it
  natively — images, and PDFs when the provider supports them), or
- a string blob to prepend to the prompt (everything else).

Imports for pypdf / openpyxl are lazy so the engine doesn't crash on
environments without the `files` extra installed; callers get a clear
"install with `uv sync --extra files`" message instead.
"""

from __future__ import annotations

import csv
import io
import mimetypes
from dataclasses import dataclass
from pathlib import Path

from hollerbox.providers.base import Attachment

# Provider-native: pass through to the LLM call as bytes.
_IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".webp", ".gif"}
_PDF_EXTS = {".pdf"}

# Extract-to-text: read locally, fold into the prompt.
_EXCEL_EXTS = {".xlsx", ".xls"}
_CSV_EXTS = {".csv", ".tsv"}
_TEXT_EXTS = {".txt", ".md", ".markdown", ".rst", ".json", ".yaml", ".yml", ".html", ".xml", ".log"}


@dataclass
class ExtractedAttachment:
    """How the LLM step ended up handling one input path."""

    name: str
    media_type: str
    # Exactly one of these is non-None:
    native_attachment: Attachment | None = None
    extracted_text: str | None = None
    # When extraction itself fails (missing optional dep, malformed file)
    # we still want to surface SOMETHING to the LLM so it can say "I
    # couldn't read X" rather than disappear.
    error: str | None = None


def _media_type_for(p: Path) -> str:
    mime, _ = mimetypes.guess_type(p.name)
    return mime or "application/octet-stream"


def extract(path: str) -> ExtractedAttachment:
    p = Path(path).expanduser()
    if not p.is_file():
        return ExtractedAttachment(
            name=p.name,
            media_type="application/octet-stream",
            error=f"file not found: {p}",
        )

    name = p.name
    suffix = p.suffix.lower()
    media_type = _media_type_for(p)

    # Image → pass straight through.
    if suffix in _IMAGE_EXTS or media_type.startswith("image/"):
        return ExtractedAttachment(
            name=name,
            media_type=media_type,
            native_attachment=Attachment(
                data=p.read_bytes(), media_type=media_type, name=name
            ),
        )

    # PDF → pass through as a native attachment AND keep an extracted-text
    # fallback. The provider decides which to use; if it can't see PDFs
    # natively it uses the text. The LLMStep is responsible for picking.
    if suffix in _PDF_EXTS or media_type == "application/pdf":
        data = p.read_bytes()
        text_fallback: str | None = None
        try:
            from pypdf import PdfReader

            reader = PdfReader(io.BytesIO(data))
            chunks: list[str] = []
            for i, page in enumerate(reader.pages):
                try:
                    chunks.append(page.extract_text() or "")
                except Exception:  # noqa: BLE001
                    chunks.append(f"[page {i + 1}: extraction failed]")
            text_fallback = "\n\n".join(chunks).strip() or None
        except ImportError:
            text_fallback = None  # provider must handle natively or fail
        except Exception as exc:  # noqa: BLE001
            text_fallback = None
            return ExtractedAttachment(
                name=name,
                media_type="application/pdf",
                native_attachment=Attachment(data=data, media_type="application/pdf", name=name),
                extracted_text=None,
                error=f"PDF text extraction failed: {exc}",
            )
        return ExtractedAttachment(
            name=name,
            media_type="application/pdf",
            native_attachment=Attachment(data=data, media_type="application/pdf", name=name),
            extracted_text=text_fallback,
        )

    # Excel → render visible cells to TSV per sheet.
    if suffix in _EXCEL_EXTS:
        try:
            from openpyxl import load_workbook  # type: ignore

            wb = load_workbook(filename=str(p), data_only=True, read_only=True)
            chunks: list[str] = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    rows.append("\t".join("" if v is None else str(v) for v in row))
                chunks.append(f"# Sheet: {sheet}\n" + "\n".join(rows))
            return ExtractedAttachment(
                name=name,
                media_type=media_type,
                extracted_text="\n\n".join(chunks),
            )
        except ImportError:
            return ExtractedAttachment(
                name=name,
                media_type=media_type,
                error=(
                    "openpyxl isn't installed — run `uv sync --extra files` to "
                    "enable Excel parsing."
                ),
            )
        except Exception as exc:  # noqa: BLE001
            return ExtractedAttachment(
                name=name,
                media_type=media_type,
                error=f"Excel extraction failed: {exc}",
            )

    # CSV / TSV → straight read.
    if suffix in _CSV_EXTS or media_type in {"text/csv", "text/tab-separated-values"}:
        try:
            with p.open(encoding="utf-8", newline="") as f:
                delim = "\t" if suffix == ".tsv" else ","
                reader = csv.reader(f, delimiter=delim)
                rows = ["\t".join(r) for r in reader]
            return ExtractedAttachment(
                name=name,
                media_type=media_type,
                extracted_text="\n".join(rows),
            )
        except Exception as exc:  # noqa: BLE001
            return ExtractedAttachment(
                name=name, media_type=media_type, error=f"CSV read failed: {exc}"
            )

    # Plain text-ish formats.
    if suffix in _TEXT_EXTS or media_type.startswith("text/"):
        try:
            return ExtractedAttachment(
                name=name,
                media_type=media_type or "text/plain",
                extracted_text=p.read_text(encoding="utf-8", errors="replace"),
            )
        except Exception as exc:  # noqa: BLE001
            return ExtractedAttachment(
                name=name, media_type=media_type, error=f"text read failed: {exc}"
            )

    # Unknown binary — refuse rather than dump bytes the LLM can't use.
    return ExtractedAttachment(
        name=name,
        media_type=media_type,
        error=(
            f"unsupported file type {suffix!r} ({media_type}). "
            "Add a workflow that converts it first."
        ),
    )
